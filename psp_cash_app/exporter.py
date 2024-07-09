import sqlite3
from openpyxl import Workbook

# Connexion à la base de données SQLite
conn = sqlite3.connect('g_notes.db')

# Création d'un objet Workbook
workbook = Workbook()

# Récupération de la liste des tables dans la base de données
cursor = conn.cursor()
cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
tables = cursor.fetchall()

# Parcours de chaque table et stockage dans une feuille Excel distincte
for table_name in tables:
    worksheet = workbook.create_sheet(title=table_name[0])
    cursor.execute(f"SELECT * FROM {table_name[0]}")
    rows = cursor.fetchall()
    for row_num, row in enumerate(rows):
        for col_num, cell_value in enumerate(row):
            worksheet.cell(row=row_num+1, column=col_num+1, value=cell_value)

# Enregistrement du fichier Excel
workbook.save('g_notes.xlsx')

# Fermeture de la connexion à la base de données
conn.close()